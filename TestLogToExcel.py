# -*- coding: UTF-8 -*-

print("\nLoading...\n")

import os
import time
import openpyxl

import tkinter as tk
from tkinter import *
from tkinter import messagebox
from tkinter.filedialog import askdirectory, askopenfilename

from openpyxl.styles import Font
from urllib.parse import unquote, urlparse
from datetime import datetime

from openpyxl.styles.colors import Color

# 宣告 "選擇資料夾" Function
def selectFolderPath():
    folderPath_ = askdirectory()
    folderPath.set(folderPath_)

# 宣告 "選擇Excel" Function
def selectExcelPath():
    excelPath_ = askopenfilename()
    excelPath.set(excelPath_)

# 宣告錯誤格式的Log，pop up message box
def pop_up():
    print("========================================\n")
    print("此為格式錯誤之檔案\n")
    messagebox.showinfo("Warning", wrongFormat)

def exit_():
    exit()

# 建立使用者 GUI，使其可選擇 資料夾 及 Excel 路徑
window = tk.Tk()
window.title("檔案選擇")
window.geometry('520x96')
window.resizable(0,0)

folderPath = StringVar()
excelPath  = StringVar()

folder_frame = tk.Frame(window)
folder_frame.grid(row = 0, column = 0, pady = 4)
folder_label = tk.Label(folder_frame, text = "資料夾路徑")
folder_label.grid(row = 0, column = 0, ipadx = 7)
folder_entry = tk.Entry(folder_frame, textvariable = folderPath)
folder_entry.grid(row = 0, column = 1, ipadx = 124)
folder_button = tk.Button(folder_frame, text = "選擇", command = selectFolderPath)
folder_button.grid(row = 0, column = 2, padx = 7)

excel_frame = tk.Frame(window)
excel_frame.grid(row = 1, column = 0, pady = 2)
excel_label = tk.Label(excel_frame, text = "Excel路徑")
excel_label.grid(row = 0, column = 0, ipadx = 10)
excel_entry = tk.Entry(excel_frame, textvariable = excelPath)
excel_entry.grid(row = 0, column = 1, ipadx = 124)
excel_button = tk.Button(excel_frame, text = "選擇", command = selectExcelPath)
excel_button.grid(row = 0, column = 2, padx = 7)

button_frame = tk.Frame(window)
button_frame.grid(row = 2, column = 0)
start_button = tk.Button(button_frame, text = "確認", command = window.destroy)
start_button.grid(row = 0, column = 0, padx = 5)
exit_button = tk.Button(button_frame, text = "關閉", command = exit_)
exit_button.grid(row = 0, column = 1, padx = 5)

print("請選擇 資料夾路徑 及 Excel 路徑\n")
print("========================================\n")
window.mainloop()

# 讀取資料夾所有檔案，放入陣列中 ( 會判別檔案形式是否正確，若錯誤則從陣列中刪除 )
allFileList = os.listdir(folderPath.get())
fileCount = len(allFileList)

print("讀取資料夾檔案中...\n")

# 排除副檔名不符合的資料
for i in range(len(allFileList)):
    if allFileList[i][-4:] != '.dcl':
        fileCount -= 1

for i in range(fileCount):
    if allFileList[i][-4:] != '.dcl':
        del allFileList[i]
        i -= 1

# 計算副檔名符合的 Log 數量
print("符合條件的 Log，有 %d 筆\n" % (fileCount))
print("========================================\n")

# 打開 Excel，進行 Initial 作業
print("打開 Excel 並進行初始化\n")

# 新增空白 Excel
newExcel = '%s.xlsx' % (datetime.now().strftime('%Y%m%d_%H%M%S'))
excelnew = openpyxl.Workbook()

# 打開範本 Excel
exceldemo = openpyxl.load_workbook(unquote(urlparse(excelPath.get()).path))
dataDemo = exceldemo['Data']
while(dataDemo.max_row > 4):
    dataDemo.delete_rows(5)
while(dataDemo.max_column > 303):
    dataDemo.delete_cols(304)
valueDemo = exceldemo['Value']
while(valueDemo.max_row > 4):
    valueDemo.delete_rows(5)
while(valueDemo.max_column > 898):
    valueDemo.delete_cols(899)
sampleDemo = exceldemo['Sample']

# 儲存空白 Excel
excelnew = exceldemo
excelnew.save(folderPath.get() + "/" + newExcel)

# 打開新建的 Excel 並開始編輯
excel = openpyxl.load_workbook(folderPath.get() + "/" + newExcel)
sheetData   = excel['Data']
sheetValue  = excel['Value']

# 建立標題行、結尾空行、運算行
indextitle = ["General Data", "", "ICT Test Date & Time", "", ""]
listtitle  = ["No.", "S/N", "Date", "Time", "Test Pass / Fail"]
end        = []
passTotal  = ["", "", "", "Total Pass"]
failTotal  = ["", "", "", "Total Fail"]
Yield      = ["", "", "", "Total Yield"]
formula    = ["", "", "", "Average"]

# 編列元件索引值
titleArray2D = []
with open(folderPath.get() + "/" + allFileList[0], 'r+') as f:
    for line in f.readlines():
        titleArray2D.append(line.split(','))

del titleArray2D[0]
del titleArray2D[0]

for i in range(len(titleArray2D)):
    indextitle.append(titleArray2D[i][0][1:])
    listtitle.append(titleArray2D[i][1][1:])

#sheetData.append(indextitle)
#sheetData.append(listtitle)

passCount   = 1
errorCount  = 0
wrongFormat = []

print("資料寫入中...\n")

# 進行 Log 資料處理，將所有資料以 "," 分開，並放進二維陣列中
for i in range(fileCount):
    global listtitleCheck, listdata, listerror, listvalue
    listtitleCheck    = {}
    listtitleCheck[i] = ["No.", "S/N", "Date", "Time", "Test Pass / Fail"]
    listdata           = {}
    listdata[i]        = []
    listerror         = {}
    listerror[i]      = []
    listvalue         = {}
    listvalue[i]      = []

    array2D = []
    with open(folderPath.get() + "/" + allFileList[i], 'r+') as f:
        print("-", end = "")
        for line in f.readlines():
            array2D.append(line.split(','))

    listdata[i].append(passCount)
    listdata[i].append(array2D[0][4])
    listdata[i].append(array2D[0][5][0:4] + "/" + array2D[0][5][4:6] + "/" + array2D[0][5][6:8])
    listdata[i].append(array2D[0][6][0:2] + ":" + array2D[0][6][2:4] + ":" + array2D[0][6][4:6])
    if array2D[0][0] == "PASS": listdata[i].append("Pass")
    else:                       listdata[i].append("Fail")

    listvalue[i].append(passCount)
    listvalue[i].append(array2D[0][4])
    listvalue[i].append(array2D[0][5][0:4] + "/" + array2D[0][5][4:6] + "/" + array2D[0][5][6:8])
    listvalue[i].append(array2D[0][6][0:2] + ":" + array2D[0][6][2:4] + ":" + array2D[0][6][4:6])
    
    del array2D[0]
    del array2D[0]

    for x in range(len(array2D)):
        if array2D[x][0][0:5] == "Short" or array2D[x][0][0:4] == "Open":
            listerror[errorCount] = array2D[-1]
            del array2D[-1]
            del array2D[-1]
            errorCount += 1
    
    for j in range(len(array2D)):
        listtitleCheck[i].append(array2D[j][1][1:])
        if array2D[j][11] == " 0\n":    listdata[i].append("Pass")
        else:                           listdata[i].append("Fail")
        listvalue[i].append(float(array2D[j][3][3:-2]))
        listvalue[i].append(float(array2D[j][9][1:-2]))
        listvalue[i].append((float(array2D[j][10][1:-2]))/100)

    if len(listerror) != -1:
        for z in range((len(listerror) - 1)):
            listdata[i].append("".join(listerror[z]))

    if listtitleCheck[i] == listtitle:
        passCount += 1
        sheetData.append(listdata[i])
        sheetValue.append(listvalue[i])
    else:
        wrongFormat.append(allFileList[i])

sheetData.append(end)
sheetData.append(passTotal)
sheetData.append(failTotal)
sheetData.append(Yield)

sheetValue.append(end)
sheetValue.append(formula)

if sheetData.max_column == 303:
    for i in range(5, (sheetData.max_column + 1)):
        celltop = sheetData.cell(row = 5, column = i)
        cellbuttom = sheetData.cell(row = passCount + 3, column = i)
        cellpass = sheetData.cell(row = passCount + 5, column = i)
        cellpass.value = "=COUNTIF({}:{}, \"Pass\")".format(celltop.coordinate, cellbuttom.coordinate)
        cellfail = sheetData.cell(row = passCount + 6, column = i)
        cellfail.value = "=COUNTIF({}:{}, \"Fail\")".format(celltop.coordinate, cellbuttom.coordinate)
        cellaverage = sheetData.cell(row = passCount + 7, column = i)
        cellaverage.value = "={}/({}+{})".format(cellpass.coordinate, cellpass.coordinate, cellfail.coordinate)
        cellaverage.number_format = '0.00%'

        passCounter = 0
        failCounter = 0
        for j in range(5, passCount + 4):
            if sheetData.cell(column = i, row = j).value == "Pass": passCounter += 1
            else:                                                   failCounter += 1
        if (passCounter + failCounter) != (passCount - 1):  print("Something Wrong!")
else:
    for i in range(5, (sheetData.max_column)):
        celltop = sheetData.cell(row = 5, column = i)
        cellbuttom = sheetData.cell(row = passCount + 3, column = i)
        cellpass = sheetData.cell(row = passCount + 5, column = i)
        cellpass.value = "=COUNTIF({}:{}, \"Pass\")".format(celltop.coordinate, cellbuttom.coordinate)
        cellfail = sheetData.cell(row = passCount + 6, column = i)
        cellfail.value = "=COUNTIF({}:{}, \"Fail\")".format(celltop.coordinate, cellbuttom.coordinate)
        cellaverage = sheetData.cell(row = passCount + 7, column = i)
        cellaverage.value = "={}/({}+{})".format(cellpass.coordinate, cellpass.coordinate, cellfail.coordinate)
        cellaverage.number_format = '0.00%'

        passCounter = 0
        failCounter = 0
        for j in range(5, passCount + 4):
            if sheetData.cell(column = i, row = j).value == "Pass": passCounter += 1
            else:                                                   failCounter += 1
        if (passCounter + failCounter) != (passCount - 1):  print("Something Wrong!")
    sheetData.cell(row = passCount + 7, column = sheetData.max_column).value = ("Short or Open : %d" % errorCount)

for i in range(5, (sheetValue.max_column + 1)):
    celltop     = sheetValue.cell(row = 5, column = i)
    cellbuttom  = sheetValue.cell(row = passCount + 3, column = i)
    cellaverage = sheetValue.cell(row = passCount + 5, column = i)
    cellaverage.value = "=AVERAGE({}:{})".format(celltop.coordinate, cellbuttom.coordinate)

for i in range(5, passCount + 4):
    for j in range(7, 899, 3):
        sheetValue.cell(row = i, column = j).number_format = '0.0%'

for i in range(7, 899, 3):
    sheetValue.cell(row = passCount + 5, column = i).number_format = '0.00%'

# 設定儲存格字型及大小
for row in sheetData:
    for cell in row:
        cell.font = Font(name = "Arial", size = 11)

sheetData['A1'].font = Font(name = "Arial", size = 14, color = 'FFFFFF', bold = True)
sheetData['A2'].font = Font(name = "Arial", size = 12, color = 'FFFFFF', bold = True)
sheetData['C2'].font = Font(name = "Arial", size = 11, color = 'FFFFFF', bold = True)
sheetData['F2'].font = Font(name = "Arial", size = 11, bold = True)
sheetData['A4'].font = Font(name = "Arial", size = 11, bold = True)
sheetData['B4'].font = Font(name = "Arial", size = 11, bold = True)
sheetData['C4'].font = Font(name = "Arial", size = 11, bold = True)
sheetData['D4'].font = Font(name = "Arial", size = 11, bold = True)
sheetData['E4'].font = Font(name = "Arial", size = 11, bold = True)

for row in sheetValue:
    for cell in row:
        cell.font = Font(name = "Arial", size = 11)

sheetValue['A1'].font = Font(name = "Arial", size = 14, color = 'FFFFFF', bold = True)
sheetValue['A2'].font = Font(name = "Arial", size = 12, color = 'FFFFFF', bold = True)
sheetValue['C2'].font = Font(name = "Arial", size = 11, color = 'FFFFFF', bold = True)
sheetValue['E2'].font = Font(name = "Arial", size = 11, bold = True)
sheetValue['A4'].font = Font(name = "Arial", size = 11, bold = True)
sheetValue['B4'].font = Font(name = "Arial", size = 11, bold = True)
sheetValue['C4'].font = Font(name = "Arial", size = 11, bold = True)
sheetValue['D4'].font = Font(name = "Arial", size = 11, bold = True)

print("\n\n資料儲存中...\n")

# 儲存編輯好的 Excel
excel.save(folderPath.get() + "/" + newExcel)

print("儲存完成！\n")
time.sleep(2)

if wrongFormat != []:   pop_up()
